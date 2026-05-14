from django.http import JsonResponse
from django.shortcuts import render
from django.db.models import Sum
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.views.decorators.cache import cache_page

def paginate_data(request, data, per_page=20):
    page_number = request.GET.get('page', 1)
    paginator = Paginator(data, per_page)
    page_obj = paginator.get_page(page_number)
    
    return {
        "results": list(page_obj),
        "total_pages": paginator.num_pages,
        "current_page": page_obj.number,
        "has_next": page_obj.has_next(),
        "has_previous": page_obj.has_previous(),
        "total_items": paginator.count
    }

from .models import Item, Sale, Purchase
import pandas as pd
import openpyxl

def safe_float(value, default=0.0):
    if pd.isna(value) or value is None or value == "":
        return default
    try:
        if isinstance(value, str):
            value = value.replace(',', '').strip()
        return float(value)
    except (ValueError, TypeError):
        return default


@csrf_exempt
def upload_page(request):
    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return JsonResponse({
                "error": "No file uploaded"
            }, status=400)
        try:
            saved_count = 0
            if uploaded_file.name.endswith('.csv'):
                chunksize = 5000
                for df in pd.read_csv(uploaded_file, encoding='utf-8-sig', on_bad_lines='skip', chunksize=chunksize):
                    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
                    sales_to_create = []
                    for _, row in df.iterrows():
                        item_code = str(row["كود الصنف"])
                        item_name = str(row["اسم الصنف"])
                        barcode = None
                        if "باركود" in df.columns and pd.notna(row["باركود"]):
                            barcode = str(row["باركود"])
                        item, created = Item.objects.get_or_create(
                            item_code=item_code,
                            defaults={
                                "item_name": item_name,
                                "barcode": barcode
                            }
                        )
                        sales_to_create.append(
                            Sale(
                                item=item,
                                quantity_sold=safe_float(row.get("صافى كمية مبيعات")),
                                sales_value=safe_float(row.get("صافى قيمة مبيعات"))
                            )
                        )
                        saved_count += 1
                    Sale.objects.bulk_create(sales_to_create, batch_size=5000)

            elif uploaded_file.name.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
                ws = wb.active
                rows = ws.iter_rows(values_only=True)
                
                header = None
                for r in rows:
                    if r and any(r):
                        header = r
                        break
                        
                if not header:
                    return JsonResponse({"error": "Empty Excel file"}, status=400)
                    
                header_map = {str(name).strip(): idx for idx, name in enumerate(header) if name}
                
                code_idx = header_map.get("كود الصنف")
                name_idx = header_map.get("اسم الصنف")
                barcode_idx = header_map.get("باركود")
                qty_idx = header_map.get("صافى كمية مبيعات")
                val_idx = header_map.get("صافى قيمة مبيعات")
                
                if code_idx is None or qty_idx is None or val_idx is None:
                    return JsonResponse({"error": "Missing required columns in Excel"}, status=400)

                sales_to_create = []
                batch_size = 5000
                
                for row in rows:
                    if not row or not row[code_idx]:
                        continue
                        
                    item_code = str(row[code_idx])
                    item_name = str(row[name_idx]) if name_idx is not None and row[name_idx] else ""
                    barcode = str(row[barcode_idx]) if barcode_idx is not None and row[barcode_idx] else None
                    
                    item, created = Item.objects.get_or_create(
                        item_code=item_code,
                        defaults={
                            "item_name": item_name,
                            "barcode": barcode
                        }
                    )
                    
                    qty = safe_float(row[qty_idx])
                    val = safe_float(row[val_idx])
                    
                    sales_to_create.append(
                        Sale(
                            item=item,
                            quantity_sold=qty,
                            sales_value=val
                        )
                    )
                    saved_count += 1
                    
                    if len(sales_to_create) >= batch_size:
                        Sale.objects.bulk_create(sales_to_create, batch_size=batch_size)
                        sales_to_create.clear()
                        
                if sales_to_create:
                    Sale.objects.bulk_create(sales_to_create, batch_size=batch_size)
                    
                wb.close()
            else:
                return JsonResponse({
                    "error": "Unsupported file type"
                }, status=400)
            
            return JsonResponse({
                "message": "Sales data uploaded successfully",
                "saved_rows": saved_count
            })
        except Exception as e:
            return JsonResponse({
                "error": str(e)
            }, status=500)

    return render(request, 'upload.html')
@csrf_exempt
def upload_purchases(request):

    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')

        if not uploaded_file:
            return JsonResponse({
                "error": "No file uploaded"
            }, status=400)
        
        try:
            saved_count = 0
            
            wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            
            for _ in range(4):
                next(rows, None)
                
            header = next(rows, None)
            if not header:
                return JsonResponse({"error": "Empty or invalid Excel file"}, status=400)
                
            header_map = {str(name).strip(): idx for idx, name in enumerate(header) if name}
            
            code_idx = header_map.get("كود الصنف")
            name_idx = header_map.get("إسم الصنف")
            qty_idx = header_map.get("صافى كمية المشتريات")
            val_idx = header_map.get("صافى المشتريات")
            
            if code_idx is None or qty_idx is None or val_idx is None:
                return JsonResponse({"error": "Missing required columns in Excel"}, status=400)

            purchases_to_create = []
            batch_size = 5000
            
            for row in rows:
                if not row or not row[code_idx]:
                    continue
                    
                if row[qty_idx] is None or row[val_idx] is None:
                    continue
                    
                item_code = str(row[code_idx])
                item_name = str(row[name_idx]) if name_idx is not None and row[name_idx] else ""
                
                item, created = Item.objects.get_or_create(
                    item_code=item_code,
                    defaults={
                        "item_name": item_name
                    }
                )
                
                qty = safe_float(row[qty_idx])
                val = safe_float(row[val_idx])
                
                purchases_to_create.append(
                    Purchase(
                        item=item,
                        quantity_purchased=qty,
                        purchase_value=val
                    )
                )
                saved_count += 1
                
                if len(purchases_to_create) >= batch_size:
                    Purchase.objects.bulk_create(purchases_to_create, batch_size=batch_size)
                    purchases_to_create.clear()
                    
            if purchases_to_create:
                Purchase.objects.bulk_create(purchases_to_create, batch_size=batch_size)
                
            wb.close()
            
            return JsonResponse({
                "message": "Purchases uploaded successfully",
                "saved_rows": saved_count
            })

        except Exception as e:
            return JsonResponse({
                "error": str(e)
            }, status=500)
    return render(request, 'upload.html')
@cache_page(60 * 5)
def sales_count(request):

    return JsonResponse({
        "items_count": Item.objects.count(),
        "sales_count": Sale.objects.count(),
        "purchases_count": Purchase.objects.count()
    })
@cache_page(60 * 5)
def top_items(request):

    top_items_data = (
        Sale.objects
        .values('item__item_name')
        .annotate(
            total_sales=Sum('sales_value')
        )
        .order_by('-total_sales')[:10]
    )

    return JsonResponse(
        list(top_items_data),
        safe=False
    )
def search_items(request):

    query = request.GET.get('q')

    if not query:

        return JsonResponse({
            "error": "Please provide search query"
        }, status=400)

    items = (
        Item.objects
        .filter(item_name__icontains=query)
        .values(
            'item_code',
            'item_name',
            'barcode'
        )
    )

    paginated_data = paginate_data(request, items, 20)

    return JsonResponse(
        paginated_data,
        safe=False
    )
@cache_page(60 * 5)
def profit_report(request):
    items = Item.objects.values('id', 'item_code', 'item_name')
    sales_qs = Sale.objects.values('item_id').annotate(total=Sum('sales_value'))
    purchases_qs = Purchase.objects.values('item_id').annotate(total=Sum('purchase_value'))
    
    sales_dict = {s['item_id']: (s['total'] or 0) for s in sales_qs}
    purchases_dict = {p['item_id']: (p['total'] or 0) for p in purchases_qs}
    
    report = []
    for item in items:
        item_id = item['id']
        t_sales = sales_dict.get(item_id, 0)
        t_purchases = purchases_dict.get(item_id, 0)
        profit = t_sales - t_purchases
        
        report.append({
            "item_code": item['item_code'],
            "item_name": item['item_name'],
            "total_sales": t_sales,
            "total_purchases": t_purchases,
            "profit": profit
        })
        
    report.sort(key=lambda x: x['profit'], reverse=True)
    paginated_data = paginate_data(request, report, 20)
    return JsonResponse(paginated_data, safe=False)
@cache_page(60 * 5)
def deadstock_items(request):
    items = Item.objects.values('id', 'item_code', 'item_name')
    sales_qs = Sale.objects.values('item_id').annotate(total=Sum('quantity_sold'))
    purchases_qs = Purchase.objects.values('item_id').annotate(total=Sum('quantity_purchased'))
    
    sales_dict = {s['item_id']: (s['total'] or 0) for s in sales_qs}
    purchases_dict = {p['item_id']: (p['total'] or 0) for p in purchases_qs}
    
    report = []
    for item in items:
        item_id = item['id']
        t_sales = sales_dict.get(item_id, 0)
        t_purchases = purchases_dict.get(item_id, 0)
        
        if t_purchases > 0 and t_sales == 0:
            report.append({
                "item_code": item['item_code'],
                "item_name": item['item_name'],
                "purchased_quantity": t_purchases,
                "sold_quantity": t_sales
            })
            
    report.sort(key=lambda x: x['purchased_quantity'], reverse=True)
    paginated_data = paginate_data(request, report, 20)
    return JsonResponse(paginated_data, safe=False)